from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import Stock, SalesRecord, SalesItem
from .forms import StockForm, SalesRecordForm, SalesItemFormSet, CentralStockForm, SaleRecordUpdateForm # Import CentralStockForm and SaleRecordUpdateForm
from .export_forms import StockExportForm, SalesExportForm
from accounts.models import Branch
from django.db.models import Q, Sum
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from django.views import View
from django.urls import reverse_lazy
from django.contrib.auth.mixins import AccessMixin
from decimal import Decimal # Import Decimal for precise calculations


@login_required
def add_stock(request):
    # Only allow admin, sales, and manager users
    if request.user.access_level not in ["admin", "manager"]:
        messages.error(request, "You do not have permission to add stock.")
        return redirect("store:stock_list")

    if request.method == "POST":
        form = StockForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            stock = form.save(commit=False)

            # For non-admins, assign branch automatically
            if request.user.access_level != "admin":
                stock.branch = request.user.branch

            stock.save()
            messages.success(request, "Stock added successfully.")
            return redirect("store:stock_list")
    else:
        form = StockForm(user=request.user)

    return render(request, "store/add_stock.html", {"form": form})


@login_required
def central_add_stock_view(request, stock_name=None):
    # Only allow admin and manager users
    if request.user.access_level not in ["admin", "manager"]:
        messages.error(request, "You do not have permission to access this page.")
        return redirect("store:stock_list")

    initial_data = {}
    if stock_name:
        # Fetch an existing stock item to get general details for pre-population
        # We assume 'name' is unique enough for a central stock item, or we'd need a UUID for the "master" stock
        existing_stock_items = Stock.objects.filter(name=stock_name)
        if existing_stock_items.exists():
            first_stock_item = existing_stock_items.first()
            initial_data["name"] = first_stock_item.name
            initial_data["unit_value"] = first_stock_item.unit_value
            initial_data["stock_item_id"] = first_stock_item.id # Pass an ID to indicate update mode

            # Pre-populate branch quantities
            for stock_item in existing_stock_items:
                field_name = f"quantity_branch_{stock_item.branch.id}"
                initial_data[field_name] = stock_item.quantity
        else:
            messages.warning(request, f"Stock item '{stock_name}' not found for editing.")
            return redirect("store:central_add_stock") # Redirect to add new if not found

    if request.method == "POST":
        form = CentralStockForm(request.POST, initial=initial_data)
        if form.is_valid():
            form.save()
            messages.success(request, "Stock updated across branches successfully.")
            return redirect("store:stock_list") # Redirect to stock list or a confirmation page
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CentralStockForm(initial=initial_data)

    branches = Branch.objects.all().order_by("name")
    branch_fields = []
    for branch in branches:
        field_name = f"quantity_branch_{branch.id}"
        if field_name in form.fields: # Ensure the field exists in the form
            branch_fields.append({
                'branch': branch,
                'field': form[field_name] # Get the BoundField
            })

    context = {
        "form": form,
        "branches": branches, # Keep branches for general info if needed
        "branch_fields": branch_fields, # New list of branch and their bound fields
        "stock_name": stock_name, # Pass stock_name to template for conditional rendering
    }
    return render(request, "store/central_add_stock.html", context)


@login_required
def stock_detail(request, pk):
    stock = get_object_or_404(Stock, pk=pk)

    if request.method == "POST":
        # Only allow admin and manager users to update stock
        if request.user.access_level not in ["admin", "manager"]:
            messages.error(request, "You do not have permission to update stock.")
            return redirect("store:stock_detail", pk=stock.pk)

        form = StockForm(request.POST, request.FILES, instance=stock, user=request.user)
        if form.is_valid():
            stock = form.save(commit=False)

            # Prevent branch change if not admin
            if request.user.access_level != "admin":
                stock.branch = request.user.branch

            stock.save()
            messages.success(request, "Stock updated successfully.")
            return redirect("store:stock_detail", pk=stock.pk)
    else:
        form = StockForm(instance=stock, user=request.user)

    return render(request, "store/stock_detail.html", {"form": form, "stock": stock})


@login_required
def stock_list(request):
    stock_items = Stock.objects.all()
    branch_filter = request.GET.get("branch")
    query = request.GET.get("q")

    # Apply branch filter if specified
    if branch_filter:
        stock_items = stock_items.filter(branch_id=branch_filter)

    if query:
        stock_items = stock_items.filter(Q(name__icontains=query))

    # Show all branches to everyone for filtering purposes
    branches = Branch.objects.all()

    from django.core.paginator import Paginator

    paginator = Paginator(stock_items, 50)  # Show 50 stock items per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "branches": branches,
        "selected_branch": branch_filter,
    }
    return render(request, "store/stock_list.html", context)


@login_required
def sales_dashboard(request):
    from_date_str = request.GET.get("from_date")
    to_date_str = request.GET.get("to_date")

    sales_records_queryset = SalesRecord.objects.all()

    if from_date_str:
        try:
            from_date = timezone.make_aware(
                datetime.strptime(from_date_str, "%Y-%m-%d")
            )
            sales_records_queryset = sales_records_queryset.filter(
                sale_date__gte=from_date
            )
        except ValueError:
            messages.error(
                request, "Invalid 'from date' format. Please use YYYY-MM-DD."
            )

    if to_date_str:
        try:
            to_date = (
                timezone.make_aware(datetime.strptime(to_date_str, "%Y-%m-%d"))
                + timedelta(days=1)
                - timedelta(microseconds=1)
            )
            sales_records_queryset = sales_records_queryset.filter(
                sale_date__lte=to_date
            )
        except ValueError:
            messages.error(request, "Invalid 'to date' format. Please use YYYY-MM-DD.")

    if request.user.access_level in ["admin", "manager"]:
        # Admin and Manager see an overview of all branches
        branches = Branch.objects.all()
        branch_sales_summary = []

        for branch in branches:
            branch_sales_qs = sales_records_queryset.filter(branch=branch)
            total_sales_records = branch_sales_qs.count()
            total_cash_sales = (
                branch_sales_qs.aggregate(Sum("amount_paid_cash"))[
                    "amount_paid_cash__sum"
                ]
                or 0
            )
            total_credit_sales = (
                branch_sales_qs.aggregate(Sum("credit_owed"))["credit_owed__sum"] or 0
            )
            total_overall_sales = (
                branch_sales_qs.aggregate(Sum("total_amount"))["total_amount__sum"] or 0
            )

            branch_sales_summary.append(
                {
                    "branch": branch,
                    "total_sales_records": total_sales_records,
                    "total_cash_sales": total_cash_sales,
                    "total_credit_sales": total_credit_sales,
                    "total_overall_sales": total_overall_sales,
                }
            )
        context = {
            "branch_sales_summary": branch_sales_summary,
            "is_admin_or_manager": True,
            "from_date": from_date_str,
            "to_date": to_date_str,
        }
        return render(request, "store/sales_overview.html", context)
    else:
        # Other staff see only their branch's sales records, redirect to sales_list_by_branch
        return redirect("store:sales_list_by_branch")


@login_required
def sales_list_by_branch(request, branch_pk=None):
    sales_records_queryset = SalesRecord.objects.all()
    branch = None

    if request.user.access_level in ["admin", "manager"]:
        if branch_pk:
            branch = get_object_or_404(Branch, pk=branch_pk)
            sales_records_queryset = sales_records_queryset.filter(branch=branch)
        else:
            # Admins/Managers viewing all sales without a specific branch filter
            pass
    else:
        # Regular sales clerks can only view their branch's sales
        user_branch = request.user.branch
        if not user_branch:
            messages.error(request, "You are not assigned to any branch.")
            return redirect("home:dashboard")
        branch = user_branch
        sales_records_queryset = sales_records_queryset.filter(branch=user_branch)

    # Categorize Sales
    category = request.GET.get("category")
    if category == "cash":
        sales_records_queryset = sales_records_queryset.filter(credit_owed=0)
    elif category == "credit":
        sales_records_queryset = sales_records_queryset.filter(credit_owed__gt=0)

    sales_records = sales_records_queryset.order_by("-sale_date")

    from django.core.paginator import Paginator
    paginator = Paginator(sales_records, 20) # Show 20 sales records per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "branch": branch,
        "selected_category": category,
        "can_create_sales": request.user.access_level == "sales",
        "is_admin_or_manager": request.user.access_level in ["admin", "manager"],
        "branches": Branch.objects.all() if request.user.access_level in ["admin", "manager"] else [], # For branch filter dropdown
    }
    return render(request, "store/sales_list.html", context)


class SalesRecordDetailView(LoginRequiredMixin, AccessMixin, View):
    template_name = "store/sales_record_detail.html"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        sales_record = get_object_or_404(SalesRecord, pk=pk)
        form = SaleRecordUpdateForm(instance=sales_record)
        context = {
            "sales_record": sales_record,
            "form": form,
            "can_edit": request.user.access_level in ["admin", "manager", "sales"],
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        sales_record = get_object_or_404(SalesRecord, pk=pk)

        # Access control for editing
        if request.user.access_level not in ["admin", "manager", "sales"]:
            messages.error(request, "You do not have permission to edit this sales record.")
            return redirect("store:sales_record_detail", pk=pk)

        form = SaleRecordUpdateForm(request.POST, instance=sales_record)
        if form.is_valid():
            # Handle "Mark as Paid" button
            if "mark_as_paid" in request.POST:
                sales_record.amount_paid_cash = sales_record.total_amount
                sales_record.credit_owed = Decimal('0.00')
                messages.success(request, "Sales record marked as paid.")
            else:
                # Update fields and recalculate credit_owed
                new_amount_paid_cash = form.cleaned_data["amount_paid_cash"]
                
                # Ensure amount_paid_cash does not exceed total_amount
                if new_amount_paid_cash > sales_record.total_amount:
                    messages.error(request, "Amount paid cannot exceed total amount.")
                    form = SaleRecordUpdateForm(instance=sales_record) # Re-initialize form with original data
                    context = {
                        "sales_record": sales_record,
                        "form": form,
                        "can_edit": request.user.access_level in ["admin", "manager", "sales"],
                    }
                    return render(request, self.template_name, context)

                sales_record.amount_paid_cash = new_amount_paid_cash
                sales_record.credit_owed = sales_record.total_amount - sales_record.amount_paid_cash
                messages.success(request, "Sales record updated successfully.")
            
            sales_record.save()
            return redirect("store:sales_record_detail", pk=pk)
        else:
            messages.error(request, "Please correct the errors below.")
        
        context = {
            "sales_record": sales_record,
            "form": form,
            "can_edit": request.user.access_level in ["admin", "manager", "sales"],
        }
        return render(request, self.template_name, context)


@login_required
def create_sales_record(request):
    # Only sales staff can create sales records
    if request.user.access_level != "sales":
        messages.error(request, "You do not have permission to create sales records.")
        return redirect("store:sales_dashboard")

    if request.method == "POST":
        form = SalesRecordForm(request.POST)
        formset = SalesItemFormSet(
            request.POST,
            instance=SalesRecord(),
            form_kwargs={"branch": request.user.branch},
        )
        if form.is_valid() and formset.is_valid():
            sales_record = form.save(commit=False)
            sales_record.branch = request.user.branch
            sales_record.save()

            total_amount = 0
            for item_form in formset:
                if item_form.cleaned_data and not item_form.cleaned_data.get("DELETE"):
                    try:
                        sales_item = item_form.save(commit=False)
                        sales_item.sales_record = sales_record
                        sales_item.save()
                        total_amount += (
                            sales_item.price_at_sale * sales_item.quantity_sold
                        )
                    except ValueError as e:
                        messages.error(request, f"Error processing sale: {e}")
                        # If there's a stock error, delete the sales record and return
                        sales_record.delete()
                        return render(
                            request,
                            "store/create_sales_record.html",
                            {"form": form, "formset": formset},
                        )
                    except Exception as e:
                        messages.error(
                            request,
                            f"Error processing sale for item: {e}",
                        )
                        sales_record.delete()
                        return render(
                            request,
                            "store/create_sales_record.html",
                            {"form": form, "formset": formset},
                        )

            sales_record.total_amount = total_amount
            sales_record.save()

            messages.success(request, "Sales record created successfully.")
            return redirect("store:sales_dashboard")
        else:
            # If form or formset is not valid, errors will be displayed in the template
            pass
    else:
        form = SalesRecordForm()
        formset = SalesItemFormSet(
            instance=SalesRecord(), form_kwargs={"branch": request.user.branch}
        )

    context = {
        "form": form,
        "formset": formset,
    }
    return render(request, "store/create_sales_record.html", context)


class StockExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    raise_exception = True  # Raise 403 if test_func returns False

    def test_func(self):
        return self.request.user.access_level in ["admin", "manager", "sales"]

    def get(self, request):
        form = StockExportForm(request.GET)
        stock_items = Stock.objects.all()

        if form.is_valid():
            start_date = form.cleaned_data.get("start_date")
            end_date = form.cleaned_data.get("end_date")
            fields_to_export = form.cleaned_data.get("fields_to_export")

            if start_date:
                stock_items = stock_items.filter(added_on__gte=start_date)
            if end_date:
                stock_items = stock_items.filter(added_on__lte=end_date)

            if "export" in request.GET:
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="stock_export.xlsx"'
                )

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Stock"

                # Apply styles
                header_font = Font(bold=True)
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                center_aligned_text = Alignment(horizontal="center")

                # Write headers
                headers = [
                    Stock._meta.get_field(field_name).verbose_name
                    for field_name in fields_to_export
                ]
                worksheet.append(headers)
                for col in range(1, len(headers) + 1):
                    worksheet.cell(row=1, column=col).font = header_font
                    worksheet.cell(row=1, column=col).border = thin_border
                    worksheet.cell(row=1, column=col).alignment = center_aligned_text

                # Write data
                for item in stock_items:
                    row_data = []
                    for field_name in fields_to_export:
                        value = getattr(item, field_name)
                        # Convert non-primitive Django model objects to their string representation
                        if hasattr(value, "__str__") and not isinstance(
                            value, (str, int, float, bool, datetime)
                        ):
                            row_data.append(str(value))
                        elif isinstance(value, datetime) and timezone.is_aware(value):
                            row_data.append(
                                timezone.make_naive(value)
                            )  # Convert to naive datetime
                        else:
                            row_data.append(value)
                    worksheet.append(row_data)
                    for col in range(1, len(row_data) + 1):
                        worksheet.cell(row=worksheet.max_row, column=col).border = (
                            thin_border
                        )

                workbook.save(response)
                return response

        # Prepare headers for template preview
        template_headers = []
        if form.is_valid() and fields_to_export:
            for field_name in fields_to_export:
                template_headers.append(Stock._meta.get_field(field_name).verbose_name)

        context = {
            "form": form,
            "stock_items": stock_items,
            "template_headers": template_headers,
        }
        return render(request, "store/stock_export.html", context)


class SalesExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    raise_exception = True  # Raise 403 if test_func returns False

    def test_func(self):
        return self.request.user.access_level in ["admin", "manager", "sales"]

    def get(self, request):
        form = SalesExportForm(request.GET)
        sales_records = SalesRecord.objects.all()

        if form.is_valid():
            start_date = form.cleaned_data.get("start_date")
            end_date = form.cleaned_data.get("end_date")
            fields_to_export = form.cleaned_data.get("fields_to_export")

            if start_date:
                sales_records = sales_records.filter(sale_date__gte=start_date)
            if end_date:
                sales_records = sales_records.filter(sale_date__lte=end_date)

            if "export" in request.GET:
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="sales_export.xlsx"'
                )

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Sales"

                # Apply styles
                header_font = Font(bold=True)
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                center_aligned_text = Alignment(horizontal="center")

                # Write headers
                headers = []
                for field_name in fields_to_export:
                    if (
                        field_name == "stock_item_name"
                    ):  # Custom field for Stock Item Name
                        headers.append("Stock Item Name")
                    elif field_name == "quantity_sold":
                        headers.append("Quantity Sold")
                    else:
                        headers.append(
                            SalesRecord._meta.get_field(field_name).verbose_name
                        )
                worksheet.append(headers)
                for col in range(1, len(headers) + 1):
                    worksheet.cell(row=1, column=col).font = header_font
                    worksheet.cell(row=1, column=col).border = thin_border
                    worksheet.cell(row=1, column=col).alignment = center_aligned_text

                # Write data
                for record in sales_records:
                    row_data = []
                    for field_name in fields_to_export:
                        if field_name == "stock_item_name":
                            sales_items = SalesItem.objects.filter(sales_record=record)
                            stock_names = ", ".join(
                                [
                                    f"{item.stock_item.name} (x{item.quantity_sold})"
                                    for item in sales_items
                                ]
                            )
                            row_data.append(stock_names)
                        elif field_name == "quantity_sold":
                            sales_items = SalesItem.objects.filter(sales_record=record)
                            total_quantity_sold = sum(
                                [item.quantity_sold for item in sales_items]
                            )
                            row_data.append(total_quantity_sold)
                        else:
                            value = getattr(record, field_name)
                            # Convert non-primitive Django model objects to their string representation
                            if hasattr(value, "__str__") and not isinstance(
                                value, (str, int, float, bool, datetime)
                            ):
                                row_data.append(str(value))
                            elif isinstance(value, datetime) and timezone.is_aware(
                                value
                            ):
                                row_data.append(
                                    timezone.make_naive(value)
                                )  # Convert to naive datetime
                            else:
                                row_data.append(value)
                    worksheet.append(row_data)
                    for col in range(1, len(row_data) + 1):
                        worksheet.cell(row=worksheet.max_row, column=col).border = (
                            thin_border
                        )

                workbook.save(response)
                return response

        # Prepare headers for template preview
        template_headers = []
        if form.is_valid() and fields_to_export:
            for field_name in fields_to_export:
                if field_name == "stock_item_name":
                    template_headers.append("Stock Item Name")
                elif field_name == "quantity_sold":
                    template_headers.append("Quantity Sold")
                else:
                    template_headers.append(
                        SalesRecord._meta.get_field(field_name).verbose_name
                    )

        context = {
            "form": form,
            "sales_records": sales_records,
            "template_headers": template_headers,
        }
        return render(request, "store/sales_export.html", context)
