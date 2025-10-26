from openpyxl import load_workbook, Workbook
import math
import os

# ======== CONFIGURATION ========
input_file = "LSM INVIOCE NEW.xlsx"  # your big Excel file
output_dir = "split_excels"         # folder to save smaller files
sheets_per_file = 500               # number of sheets per output file
# =================================

os.makedirs(output_dir, exist_ok=True)

# Load the workbook (read-only mode is safer for large files)
print("🔄 Loading workbook... This might take a few minutes.")
wb = load_workbook(input_file, read_only=False, data_only=True)
sheet_names = wb.sheetnames
total_sheets = len(sheet_names)

print(f"✅ Total sheets found: {total_sheets}")

# Calculate number of output files
num_files = math.ceil(total_sheets / sheets_per_file)
print(f"📂 Splitting into {num_files} smaller files...")

for i in range(num_files):
    start = i * sheets_per_file
    end = min((i + 1) * sheets_per_file, total_sheets)
    subset = sheet_names[start:end]
    
    new_wb = Workbook()
    # remove default sheet
    if "Sheet" in new_wb.sheetnames:
        del new_wb["Sheet"]
    
    for sheet_name in subset:
        ws_original = wb[sheet_name]
        ws_new = new_wb.create_sheet(title=sheet_name)
        
        # Copy cell values (not formulas)
        for row in ws_original.iter_rows(values_only=True):
            ws_new.append(row)
    
    output_file = os.path.join(output_dir, f"part_{i+1}.xlsx")
    new_wb.save(output_file)
    print(f"✅ Saved {output_file} ({len(subset)} sheets)")

print("\n🎯 Splitting complete!")
