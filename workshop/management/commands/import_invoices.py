import uuid
import gc
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.core.management.base import BaseCommand
from django.db import connection
from openpyxl import load_workbook
from workshop.models import Vehicle, InternalEstimate, EstimatePart, Branch


class Command(BaseCommand):
    help = "Import vehicle, internal estimate, and parts data from Excel file (each sheet = one vehicle record)"

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        file_path = options["excel_file"]

        # ✅ Ensure 'Abuja' branch exists or create it
        branch, _ = Branch.objects.get_or_create(name="Abuja")

        total_vehicles = 0
        total_estimates = 0
        total_parts = 0
        skipped_sheets = []

        self.stdout.write(self.style.WARNING(f"🚀 Starting import from: {file_path}\n"))

        # ✅ Load workbook metadata only once
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            try:
                self.stdout.write(self.style.NOTICE(f"📄 Processing sheet: {sheet_name}"))

                ws = wb[sheet_name]

                # ✅ Extract main vehicle data
                customer_name = ws["C17"].value or ""
                address = ws["C18"].value or ""
                phone = ws["C20"].value or ""
                vehicle_make = ws["H17"].value or ""
                job_no = ws["C21"].value or ""
                model = ws["H18"].value or ""
                year = ws["H19"].value or 0
                chasis_no = ws["H20"].value or ""
                licence_plate = ws["H21"].value or ""
                mileage = ws["H22"].value or ""
                date_str = ws["C22"].value or ""

                # ✅ Skip if essential info is missing
                if not all([customer_name, vehicle_make, model, licence_plate]):
                    skipped_sheets.append(sheet_name)
                    self.stdout.write(
                        self.style.WARNING(f"⚠️ Skipped '{sheet_name}' — missing essential vehicle data")
                    )
                    continue

                # ✅ Parse registration date safely
                try:
                    date_of_first_registration = (
                        datetime.strptime(str(date_str), "%d/%m/%Y").date()
                        if date_str else datetime.now().date()
                    )
                except Exception:
                    date_of_first_registration = datetime.now().date()

                # ✅ Create Vehicle
                vehicle = Vehicle.objects.create(
                    uuid=str(uuid.uuid4())[:12],
                    branch=branch,
                    customer_name=customer_name,
                    address=address,
                    phone=phone,
                    job_no=job_no,
                    vehicle_make=vehicle_make,
                    model=model,
                    year=int(year) if year else 0,
                    chasis_no=chasis_no,
                    licence_plate=licence_plate,
                    date_of_first_registration=date_of_first_registration,
                    mileage=mileage,
                    complaint=" ",
                    status="completed",
                )
                total_vehicles += 1

                # ✅ Create Internal Estimate
                estimate = InternalEstimate.objects.create(
                    vehicle=vehicle,
                    apply_vat=False,
                    is_invoice=False,
                    vat_amount=Decimal("0.00"),
                    total_with_vat=Decimal("0.00"),
                )
                total_estimates += 1

                # ✅ Process job parts
                start_row = 25
                parts_count = 0
                parts_total = Decimal("0.00")

                while True:
                    desc = ws[f"B{start_row}"].value
                    if not desc:
                        start_row += 1
                        continue
                    if str(desc).strip().upper().startswith("SUB-TOTAL"):
                        break

                    qty = ws[f"E{start_row}"].value
                    total_amount = ws[f"G{start_row}"].value

                    # ✅ Safe quantity parsing
                    try:
                        quantity = int(qty) if qty not in (None, "", 0) else 1
                    except (TypeError, ValueError):
                        quantity = 1

                    # ✅ Parse total amount safely
                    try:
                        if isinstance(total_amount, (int, float)):
                            total_amount = Decimal(str(round(total_amount, 2)))
                        elif isinstance(total_amount, str):
                            total_amount = Decimal(total_amount.replace(",", "").strip() or "0")
                        else:
                            total_amount = Decimal("0.00")
                    except (InvalidOperation, TypeError, ValueError):
                        total_amount = Decimal("0.00")

                    # ✅ Calculate unit price
                    try:
                        price = (total_amount / quantity).quantize(Decimal("0.01")) if quantity > 0 else Decimal("0.00")
                    except (ArithmeticError, TypeError, InvalidOperation):
                        price = Decimal("0.00")

                    EstimatePart.objects.create(
                        estimate=estimate,
                        name=str(desc).strip(),
                        quantity=quantity,
                        price=price,
                    )

                    parts_total += total_amount
                    total_parts += 1
                    parts_count += 1
                    start_row += 1

                # ✅ VAT handling
                vat_row = start_row + 1
                vat_value = ws[f"H{vat_row}"].value
                try:
                    if vat_value and "7.5" in str(vat_value):
                        estimate.apply_vat = True
                        estimate.vat_amount = (parts_total * Decimal("0.075")).quantize(Decimal("0.01"))
                        estimate.total_with_vat = (parts_total + estimate.vat_amount).quantize(Decimal("0.01"))
                    else:
                        estimate.apply_vat = False
                        estimate.vat_amount = Decimal("0.00")
                        estimate.total_with_vat = parts_total.quantize(Decimal("0.01"))
                    estimate.save()
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"VAT calculation error in '{sheet_name}': {e}"))

                self.stdout.write(
                    self.style.SUCCESS(
                        f"✅ Imported {vehicle.customer_name} ({vehicle.vehicle_make} - {vehicle.licence_plate}) — {parts_count} parts linked"
                    )
                )

                # ✅ Free up memory after processing each sheet
                del ws
                connection.close()
                gc.collect()

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Error in sheet '{sheet_name}': {e}"))
                continue

        wb.close()

        # ✅ Summary
        self.stdout.write(self.style.SUCCESS(
            f"\n🎯 Import complete!\n"
            f"Vehicles imported: {total_vehicles}\n"
            f"Internal Estimates: {total_estimates}\n"
            f"Estimate Parts: {total_parts}\n"
        ))

        if skipped_sheets:
            self.stdout.write(self.style.WARNING(f"⚠️ Skipped sheets: {', '.join(skipped_sheets)}"))
