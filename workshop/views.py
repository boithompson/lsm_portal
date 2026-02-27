from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views import View
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from .models import Vehicle, InternalEstimate, EstimatePart, JobSheet, VehicleStatus
from .export_forms import WorkshopExportForm
from accounts.models import Branch, CustomUser
from django.db.models import Q, Max
from django.contrib import messages
from django.forms import inlineformset_factory
from django.urls import reverse_lazy
from django.contrib.auth.mixins import AccessMixin
from decimal import Decimal
import datetime  # Import datetime
from functools import wraps
import logging

logger = logging.getLogger(__name__)


def workshop_access_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("accounts:login")  # Redirect to login if not authenticated
        if request.user.access_level not in [
            "admin",
            "workshop",
            "procurement",
            "account",
        ]:
            messages.error(request, "You do not have permission to access this page.")
            return redirect("home:dashboard")
        return view_func(request, *args, **kwargs)

    return _wrapped_view


@login_required
@workshop_access_required
def print_proforma_invoice(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    internal_estimate = get_object_or_404(InternalEstimate, vehicle=vehicle)
    estimate_parts = internal_estimate.estimatepart_set.all()

    # The apply_vat field should be set when the InternalEstimate is created/edited.
    # We will just use the value stored in the database.
    # Ensure the save method is called to update VAT if parts have changed
    internal_estimate.save()

    # Calculate subtotal
    subtotal = sum(part.price * part.quantity for part in estimate_parts)

    # Handle date selection
    date_selected = request.GET.get("date", "now")  # Default to 'now'

    context = {
        "vehicle": vehicle,
        "internal_estimate": internal_estimate,
        "estimate_parts": estimate_parts,
        "subtotal": subtotal,
        "date_now": datetime.date.today(),  # Add current date to context
        "date_selected": date_selected,  # Pass selected date option to template
        "vat_applied": internal_estimate.apply_vat,
        "is_invoice": internal_estimate.is_invoice,  # Pass invoice status to template
    }
    return render(request, "workshop/proforma_invoice.html", context)


class WorkshopExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    raise_exception = True  # Raise 403 if test_func returns False

    def test_func(self):
        return self.request.user.access_level in ["admin", "manager", "workshop"]

    def get(self, request):
        form = WorkshopExportForm(request.GET)
        vehicles = Vehicle.objects.all()

        if form.is_valid():
            start_date = form.cleaned_data.get("start_date")
            end_date = form.cleaned_data.get("end_date")
            fields_to_export = form.cleaned_data.get("fields_to_export")

            if start_date:
                vehicles = vehicles.filter(date_created__gte=start_date)
            if end_date:
                vehicles = vehicles.filter(date_created__lte=end_date)

            if "export" in request.GET:
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="workshop_export.xlsx"'
                )

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Workshop Records"

                # Apply styles
                header_font = Font(bold=True)
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                center_aligned_text = Alignment(horizontal="center")

                # Write headers
                headers = []
                for field_name in fields_to_export:
                    if field_name in [
                        "job_sheet_service_advisor",
                        "job_sheet_assigned_to",
                        "job_sheet_accessories",
                        "job_sheet_job_description",
                    ]:
                        headers.append(
                            {
                                "job_sheet_service_advisor": "Service Advisor",
                                "job_sheet_assigned_to": "Assigned To",
                                "job_sheet_accessories": "Accessories",
                                "job_sheet_job_description": "Job Description",
                            }[field_name]
                        )
                    elif field_name in [
                        "estimate_apply_vat",
                        "estimate_discount_amount",
                        "estimate_is_invoice",
                        "estimate_grand_total",
                        "estimate_parts_details",
                    ]:
                        headers.append(
                            {
                                "estimate_apply_vat": "Apply VAT",
                                "estimate_discount_amount": "Discount Amount",
                                "estimate_is_invoice": "Is Invoice",
                                "estimate_grand_total": "Grand Total",
                                "estimate_parts_details": "Parts Details",
                            }[field_name]
                        )
                    else:
                        headers.append(Vehicle._meta.get_field(field_name).verbose_name)
                worksheet.append(headers)
                for col in range(1, len(headers) + 1):
                    worksheet.cell(row=1, column=col).font = header_font
                    worksheet.cell(row=1, column=col).border = thin_border
                    worksheet.cell(row=1, column=col).alignment = center_aligned_text

                # Write data
                for vehicle in vehicles:
                    row_data = []
                    for field_name in fields_to_export:
                        if field_name in [
                            "job_sheet_service_advisor",
                            "job_sheet_assigned_to",
                            "job_sheet_accessories",
                            "job_sheet_job_description",
                        ]:
                            try:
                                job_sheet = JobSheet.objects.get(vehicle=vehicle)
                                value = getattr(
                                    job_sheet, field_name.replace("job_sheet_", "")
                                )
                                row_data.append(str(value) if value else "")
                            except JobSheet.DoesNotExist:
                                row_data.append("")
                        elif field_name in [
                            "estimate_apply_vat",
                            "estimate_discount_amount",
                            "estimate_is_invoice",
                            "estimate_grand_total",
                        ]:
                            try:
                                estimate = InternalEstimate.objects.get(vehicle=vehicle)
                                if field_name == "estimate_grand_total":
                                    value = estimate.grand_total
                                else:
                                    value = getattr(
                                        estimate, field_name.replace("estimate_", "")
                                    )
                                row_data.append(str(value) if value else "")
                            except InternalEstimate.DoesNotExist:
                                row_data.append("")
                        elif field_name == "estimate_parts_details":
                            try:
                                estimate = InternalEstimate.objects.get(vehicle=vehicle)
                                parts = EstimatePart.objects.filter(estimate=estimate)
                                parts_details = "; ".join(
                                    [
                                        f"{part.name} (x{part.quantity}) - {part.price}"
                                        for part in parts
                                    ]
                                )
                                row_data.append(parts_details)
                            except InternalEstimate.DoesNotExist:
                                row_data.append("")
                        else:
                            value = getattr(vehicle, field_name)
                            # Convert non-primitive Django model objects to their string representation
                            if hasattr(value, "__str__") and not isinstance(
                                value,
                                (
                                    str,
                                    int,
                                    float,
                                    bool,
                                    datetime.date,
                                    datetime.datetime,
                                ),
                            ):
                                row_data.append(str(value))
                            elif isinstance(
                                value, datetime.datetime
                            ) and timezone.is_aware(value):
                                row_data.append(
                                    timezone.make_naive(value)
                                )  # Convert to naive datetime
                            else:
                                row_data.append(value)
                    worksheet.append(row_data)
                    for col in range(1, len(row_data) + 1):
                        worksheet.cell(row=worksheet.max_row, column=col).border = (
                            thin_border
                        )

                workbook.save(response)
                return response

        # Prepare headers for template preview
        template_headers = []
        if form.is_valid() and fields_to_export:
            for field_name in fields_to_export:
                if field_name in [
                    "job_sheet_service_advisor",
                    "job_sheet_assigned_to",
                    "job_sheet_accessories",
                    "job_sheet_job_description",
                ]:
                    template_headers.append(
                        {
                            "job_sheet_service_advisor": "Service Advisor",
                            "job_sheet_assigned_to": "Assigned To",
                            "job_sheet_accessories": "Accessories",
                            "job_sheet_job_description": "Job Description",
                        }[field_name]
                    )
                elif field_name in [
                    "estimate_apply_vat",
                    "estimate_discount_amount",
                    "estimate_is_invoice",
                    "estimate_grand_total",
                    "estimate_parts_details",
                ]:
                    template_headers.append(
                        {
                            "estimate_apply_vat": "Apply VAT",
                            "estimate_discount_amount": "Discount Amount",
                            "estimate_is_invoice": "Is Invoice",
                            "estimate_grand_total": "Grand Total",
                            "estimate_parts_details": "Parts Details",
                        }[field_name]
                    )
                else:
                    template_headers.append(
                        Vehicle._meta.get_field(field_name).verbose_name
                    )

        # Prepare vehicle data for template with all related information
        vehicle_data = []
        for vehicle in vehicles[:20]:  # Limit to first 20 for preview
            vehicle_row = {}
            for field_name in fields_to_export:
                if field_name in [
                    "job_sheet_service_advisor",
                    "job_sheet_assigned_to",
                    "job_sheet_accessories",
                    "job_sheet_job_description",
                ]:
                    try:
                        job_sheet = JobSheet.objects.get(vehicle=vehicle)
                        field_suffix = field_name.replace("job_sheet_", "")
                        value = getattr(job_sheet, field_suffix, "")
                        vehicle_row[field_name] = str(value) if value else ""
                    except JobSheet.DoesNotExist:
                        vehicle_row[field_name] = ""
                elif field_name in [
                    "estimate_apply_vat",
                    "estimate_discount_amount",
                    "estimate_is_invoice",
                    "estimate_grand_total",
                ]:
                    try:
                        estimate = InternalEstimate.objects.get(vehicle=vehicle)
                        field_suffix = field_name.replace("estimate_", "")
                        if field_suffix == "grand_total":
                            value = estimate.grand_total
                        else:
                            value = getattr(estimate, field_suffix, "")
                        vehicle_row[field_name] = str(value) if value else ""
                    except InternalEstimate.DoesNotExist:
                        vehicle_row[field_name] = ""
                elif field_name == "estimate_parts_details":
                    try:
                        estimate = InternalEstimate.objects.get(vehicle=vehicle)
                        parts = EstimatePart.objects.filter(estimate=estimate)
                        parts_details = "; ".join(
                            [
                                f"{part.name} (x{part.quantity}) - {part.price}"
                                for part in parts
                            ]
                        )
                        vehicle_row[field_name] = parts_details
                    except InternalEstimate.DoesNotExist:
                        vehicle_row[field_name] = ""
                else:
                    value = getattr(vehicle, field_name, "")
                    if field_name == "branch":
                        vehicle_row[field_name] = (
                            vehicle.branch.name if vehicle.branch else ""
                        )
                    elif field_name == "status":
                        vehicle_row[field_name] = vehicle.get_status_display()
                    else:
                        vehicle_row[field_name] = str(value) if value else ""
            vehicle_data.append(vehicle_row)

        context = {
            "form": form,
            "vehicles": vehicles,
            "template_headers": template_headers,
            "vehicle_data": vehicle_data,
            "fields_to_export": fields_to_export,
        }
        return render(request, "workshop/workshop_export.html", context)
